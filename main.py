import requests
import json
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
import get_teritory
import get_city
import get_party_type
import get_potential
import get_products
import get_category

url = "https://sfa3.mankindpharma.in/wsfaapi/api/newparty/saveChemist"

root = tk.Tk()
root.withdraw()
file_path = filedialog.askopenfilename()


wb = load_workbook(file_path)
ws = wb["Sheet1"]

teritories = get_teritory.run()
cities = get_city.run()
party_types = get_party_type.run()
potentials = get_potential.run()
products = get_products.run()
categories = get_category.run()

def get_teritory_code(teritory_name):
  try:
    terotory_code = teritories[teritory_name]
    return terotory_code
  except:
    raise Exception(f"Code not found for a Teritory named {teritory_name}")

def get_city_code(city_name):
  try:
    citie_code = cities[city_name]
    return citie_code
  except:
    raise Exception(f"Code not found for a City named {city_name}")

def get_party_type_code(patry_type_name):
  try:
    patry_type_code = party_types[patry_type_name]
    return patry_type_code
  except:
    raise Exception(f"Code not found for a Party type named {patry_type_name}")

def get_potential_code(potential_name):
  try:
    potential_code = potentials[potential_name]
    return potential_code
  except:
    raise Exception(f"Code not found for a Potential named {potential_name}")

def get_product_code(product_name):
  try:
    product_code = products[product_name]
    return product_code
  except:
    raise Exception(f"'Code not found for a Product named {product_name}'")


def get_categegory_code(category_name):
  category_codes = {'0-5000' : 'D', '5001-20000' : 'C','20001-50000' : 'B','50001-100000':'A','100001-ABOVE':'A+'}
  try:
    category_code = categories[category_codes[category_name]]
    return category_code
  except:
    raise Exception(f"Code not found for a Category named {category_name}")


if __name__ == '__main__':
  n = 2
  while n < 40:
    if ws.cell(row=n, column=1).value == '' or ws.cell(row=n, column=1).value == None:
      break
    payload = json.dumps([
    {
      "PARTYCODE": None,
      "PARTYNAME": ws.cell(row=n, column=1).value,
      "TYPECODE": get_party_type_code(ws.cell(row=n, column=10).value),
      "CATGCODE": get_categegory_code(ws.cell(row=n, column=11).value),
      "AREACODE": get_teritory_code(ws.cell(row=n, column=9).value),
      "CITYCODE": get_city_code(ws.cell(row=n, column=8).value),
      "ADDRESS": ws.cell(row=n, column=2).value,
      "PINCODE": ws.cell(row=n, column=4).value,
      "LANDMARK": ws.cell(row=n, column=3).value,
      "MOBILE": ws.cell(row=n, column=5).value,
      "EMAIL": "",
      "POTCODE": get_potential_code(ws.cell(row=n, column=11).value),
      "SUPPORT": ws.cell(row=n, column=7).value,
      "FREQUENCY": "2",
      "STAFFCODE": "109320",
      "STATUS": "1",
      "DIVCODE": "96",
      "CONTACT": ws.cell(row=n, column=6).value,
      "RXCODE1": get_product_code(ws.cell(row=n, column=12).value),
      "RXSUPORT1": ws.cell(row=n, column=13).value,
      "RXCODE2": "",
      "RXSUPORT2": "",
      "RXCODE3": "",
      "RXSUPORT3": "",
      "RXCODE4": "",
      "RXSUPORT4": "",
      "REMARKS": ""
    }
    ])
    print(payload)
    n += 1
    headers = {
      'Authorization': 'Basic MTAwMzkwNTE6MTA5MzIw',
      'Content-Type': 'application/json'
    }
    #response = requests.request("POST", url, headers=headers, data=payload)
    #print(response.text)


